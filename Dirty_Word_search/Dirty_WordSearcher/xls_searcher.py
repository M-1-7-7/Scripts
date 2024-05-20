import os
import re
from collections import defaultdict
import openpyxl
import xlrd

# Function to find Excel files
def find_excel_files(directory):
    """
    This function finds all Excel spreadsheets (files with .xls or .xlsx extensions)
    within the specified directory and its subdirectories.

    Args:
        directory (str): Path to the directory to search.

    Returns:
        list: List containing the full paths to all Excel spreadsheets found.
    """
    excel_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(('.xls', '.xlsx')):
                excel_files.append(os.path.join(root, file))
    return excel_files

# Function to load dirty words from file
def load_dirty_words(filename):
    """
    This function loads dirty words from a text file, removing any leading/trailing spaces
    and handling empty lines or an empty file gracefully.

    Args:
        filename (str): Path to the text file containing dirty words.

    Returns:
        list: List of cleaned dirty words (no leading/trailing spaces) or an empty list
              if the file is empty or contains only empty lines.
    """
    dirty_words = []
    try:
        with open(filename, 'r') as file:
            for line in file:
                # Remove leading/trailing spaces and add to list (if not empty)
                cleaned_word = line.strip()
                if cleaned_word:
                    dirty_words.append(cleaned_word.lower())
    except FileNotFoundError:
        print("Error: 'Dirty_word.txt' not found. Please ensure the file exists.")
    return dirty_words

# Load Dirty Words file
dirty_words = load_dirty_words('Dirty_word.txt')
if not dirty_words:
    print("Warning: 'Dirty_word.txt' is empty or contains only empty lines. No dirty words found.")

# Specify directory path (consider user input or relative paths)
computer_root = "C:\\"  # Replace with desired directory

excel_files = find_excel_files(computer_root)

# Dictionary to store findings for summary
findings_summary = defaultdict(lambda: defaultdict(int))

def process_xlsx(file_path):
    """
    Process .xlsx files to find dirty words in all cells and comments.

    Args:
        file_path (str): Path to the .xlsx file to process.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Error: Invalid .xlsx file {file_path}")
        return
    except Exception as e:
        print(f"Error opening .xlsx file {file_path}: {e}")
        return

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        # Process cell text
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_text = cell.value.lower()
                    for word in dirty_words:
                        word_pattern = r'\b' + re.escape(word) + r'\b'
                        matches = re.findall(word_pattern, cell_text)
                        if matches:
                            findings_summary[(file_path, 'Cell Text')][word] += len(matches)
                # Process comments
                if cell.comment and cell.comment.text:
                    comment_text = cell.comment.text.lower()
                    for word in dirty_words:
                        word_pattern = r'\b' + re.escape(word) + r'\b'
                        matches = re.findall(word_pattern, comment_text)
                        if matches:
                            findings_summary[(file_path, 'Comments')][word] += len(matches)

def process_xls(file_path):
    """
    Process .xls files to find dirty words in all cells.

    Args:
        file_path (str): Path to the .xls file to process.
    """
    try:
        workbook = xlrd.open_workbook(file_path)
    except xlrd.biffh.XLRDError:
        print(f"Error: Invalid .xls file {file_path}")
        return
    except Exception as e:
        print(f"Error opening .xls file {file_path}: {e}")
        return

    for sheet in workbook.sheets():
        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                cell_value = sheet.cell_value(row, col)
                if isinstance(cell_value, str):
                    cell_text = cell_value.lower()
                    for word in dirty_words:
                        word_pattern = r'\b' + re.escape(word) + r'\b'
                        matches = re.findall(word_pattern, cell_text)
                        if matches:
                            findings_summary[(file_path, 'Cell Text')][word] += len(matches)

for excel_file in excel_files:
    try:
        if excel_file.lower().endswith('.xlsx'):
            process_xlsx(excel_file)
        elif excel_file.lower().endswith('.xls'):
            process_xls(excel_file)
    except PermissionError:
        print(f"Permission error processing {excel_file}")
    except Exception as e:
        print(f"Unexpected error processing {excel_file}: {e}")

# Print summary of findings
for (file, source), words in findings_summary.items():
    words_summary = ', '.join([f"{word}: {count}" for word, count in words.items()])
    print(f"File: {file} - {source} - Contains dirty words - {words_summary}")
